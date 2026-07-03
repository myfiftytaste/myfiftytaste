"""
Inspect data/raw/Megabank.xlsx without modifying it.

The script generates:
- data/audit/megabank_inspection_report.md
- data/audit/megabank_columns.json
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `pip install openpyxl`."
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = PROJECT_ROOT / "data" / "raw" / "Megabank.xlsx"
AUDIT_DIR = PROJECT_ROOT / "data" / "audit"
REPORT_PATH = AUDIT_DIR / "megabank_inspection_report.md"
JSON_PATH = AUDIT_DIR / "megabank_columns.json"

MAX_SAMPLE_VALUES = 200
FIRST_USEFUL_ROWS = 5


CANDIDATE_RULES = {
    "title": {
        "header": ["title", "film", "movie", "name", "titre"],
        "negative": ["original", "url", "link"],
    },
    "year": {
        "header": ["year", "release year", "annee", "date"],
        "negative": ["watched", "watch date"],
    },
    "letterboxd_url": {
        "header": ["letterboxd", "url", "uri", "link", "href"],
        "negative": [],
    },
    "letterboxd_average": {
        "header": [
            "average",
            "avg",
            "rating",
            "letterboxd average",
            "note",
            "score",
        ],
        "negative": ["user", "my", "your", "count", "votes"],
    },
    "watches": {
        "header": ["watch", "watches", "watched", "views", "viewings"],
        "negative": ["date", "url"],
    },
    "likes": {
        "header": ["like", "likes"],
        "negative": [],
    },
    "fans": {
        "header": ["fan", "fans"],
        "negative": [],
    },
    "genres": {
        "header": ["genre", "genres"],
        "negative": [],
    },
    "director": {
        "header": ["director", "directors", "realisateur", "realisatrice"],
        "negative": [],
    },
    "runtime": {
        "header": ["runtime", "duration", "length", "minutes", "duree"],
        "negative": [],
    },
    "country": {
        "header": ["country", "countries", "nation", "pays"],
        "negative": [],
    },
    "language": {
        "header": ["language", "languages", "langue"],
        "negative": [],
    },
}

ROLE_LABELS = {
    "title": "titre du film",
    "year": "annee",
    "letterboxd_url": "URL Letterboxd",
    "letterboxd_average": "moyenne Letterboxd",
    "watches": "nombre de watches",
    "likes": "nombre de likes",
    "fans": "nombre de fans",
    "genres": "genres",
    "director": "realisateur",
    "runtime": "duree",
    "country": "pays",
    "language": "langue",
}


def is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def normalize_name(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_title(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_url(value: Any) -> str:
    text = clean_text(value).lower()
    text = text.split("?")[0].split("#")[0].rstrip("/")
    return text


def markdown_escape(value: Any) -> str:
    text = clean_text(value)
    return text.replace("|", "\\|").replace("\n", " ")


def make_header(raw_header: Any, index: int) -> dict[str, Any]:
    letter = get_column_letter(index)
    header = clean_text(raw_header)
    return {
        "index": index,
        "letter": letter,
        "header": header,
        "label": header if header else f"(empty header {letter})",
        "normalized": normalize_name(header),
    }


def infer_value_kind(value: Any) -> str:
    if is_empty(value):
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date)):
        return "date"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "integer" if value.is_integer() else "decimal"
    text = clean_text(value)
    if re.fullmatch(r"-?\d+", text):
        return "integer"
    if re.fullmatch(r"-?\d+[.,]\d+", text):
        return "decimal"
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text) or re.fullmatch(
        r"\d{1,2}/\d{1,2}/\d{2,4}", text
    ):
        return "date-like text"
    if text.lower() in {"true", "false", "yes", "no", "oui", "non"}:
        return "boolean-like text"
    if text.startswith(("http://", "https://", "www.")):
        return "url"
    return "text"


def infer_column_type(values: list[Any]) -> str:
    kinds = Counter(infer_value_kind(value) for value in values if not is_empty(value))
    if not kinds:
        return "empty"
    if set(kinds).issubset({"integer"}):
        return "integer"
    if set(kinds).issubset({"integer", "decimal"}):
        return "numeric"
    if set(kinds).issubset({"date", "date-like text"}):
        return "date"
    if set(kinds).issubset({"boolean", "boolean-like text"}):
        return "boolean"
    if kinds["url"] and kinds["url"] / sum(kinds.values()) >= 0.8:
        return "url"
    if kinds["text"] / sum(kinds.values()) >= 0.8:
        return "text"
    return "mixed"


def sample_score_for_role(role: str, values: list[Any]) -> float:
    non_empty = [value for value in values if not is_empty(value)]
    if not non_empty:
        return 0.0

    checked = non_empty[:MAX_SAMPLE_VALUES]
    texts = [clean_text(value).lower() for value in checked]

    if role == "letterboxd_url":
        return sum("letterboxd.com" in text for text in texts) / len(checked) * 6
    if role == "year":
        matches = 0
        for value in checked:
            text = clean_text(value)
            year = None
            if isinstance(value, (int, float)):
                year = int(value)
            elif re.fullmatch(r"\d{4}", text):
                year = int(text)
            if year and 1870 <= year <= 2035:
                matches += 1
        return matches / len(checked) * 5
    if role == "letterboxd_average":
        matches = 0
        for value in checked:
            text = clean_text(value).replace(",", ".")
            try:
                number = float(text)
            except ValueError:
                continue
            if 0 <= number <= 5:
                matches += 1
        return matches / len(checked) * 3
    if role in {"watches", "likes", "fans", "runtime"}:
        matches = 0
        for value in checked:
            text = clean_text(value).replace(",", "")
            if re.fullmatch(r"\d+(\.0)?", text):
                matches += 1
        return matches / len(checked) * 2
    if role in {"genres", "director", "country", "language"}:
        separators = sum(("," in text or ";" in text or "|" in text) for text in texts)
        return min(1.5, separators / len(checked) * 2)
    return 0.0


def header_score_for_role(role: str, header: dict[str, Any]) -> float:
    normalized = header["normalized"]
    if not normalized:
        return 0.0

    rule = CANDIDATE_RULES[role]
    score = 0.0
    for term in rule["header"]:
        normalized_term = normalize_name(term)
        if normalized == normalized_term:
            score += 5.0
        elif normalized_term in normalized:
            score += 2.5
    for term in rule["negative"]:
        if normalize_name(term) in normalized:
            score -= 2.0
    return score


def find_candidate_columns(
    headers: list[dict[str, Any]], column_values: dict[int, list[Any]]
) -> dict[str, list[dict[str, Any]]]:
    candidates: dict[str, list[dict[str, Any]]] = {}
    for role in CANDIDATE_RULES:
        scored = []
        for header in headers:
            score = header_score_for_role(role, header) + sample_score_for_role(
                role, column_values[header["index"]]
            )
            if score > 0:
                scored.append(
                    {
                        "column": header["label"],
                        "letter": header["letter"],
                        "score": round(score, 2),
                    }
                )
        candidates[role] = sorted(scored, key=lambda item: item["score"], reverse=True)[:5]
    return candidates


def best_candidate(candidates: dict[str, list[dict[str, Any]]], role: str) -> str | None:
    if not candidates.get(role):
        return None
    return candidates[role][0]["letter"]


def detect_duplicates(
    useful_rows: list[dict[str, Any]], candidates: dict[str, list[dict[str, Any]]]
) -> dict[str, Any]:
    url_letter = best_candidate(candidates, "letterboxd_url")
    title_letter = best_candidate(candidates, "title")
    year_letter = best_candidate(candidates, "year")

    result: dict[str, Any] = {
        "by_letterboxd_url": {"candidate_column": url_letter, "duplicate_groups": []},
        "by_normalized_title_year": {
            "candidate_columns": {"title": title_letter, "year": year_letter},
            "duplicate_groups": [],
        },
    }

    if url_letter:
        url_groups: dict[str, list[int]] = defaultdict(list)
        for row in useful_rows:
            row_map = {cell["letter"]: cell["value"] for cell in row["cells"]}
            url = normalize_url(row_map.get(url_letter))
            if url and "letterboxd.com" in url:
                url_groups[url].append(row["excel_row"])
        result["by_letterboxd_url"]["duplicate_groups"] = [
            {"value": value, "rows": rows}
            for value, rows in sorted(url_groups.items())
            if len(rows) > 1
        ]

    if title_letter and year_letter:
        title_year_groups: dict[str, list[int]] = defaultdict(list)
        for row in useful_rows:
            row_map = {cell["letter"]: cell["value"] for cell in row["cells"]}
            title = normalize_title(row_map.get(title_letter))
            year = clean_text(row_map.get(year_letter))
            year_match = re.search(r"\d{4}", year)
            if title and year_match:
                key = f"{title} ({year_match.group(0)})"
                title_year_groups[key].append(row["excel_row"])
        result["by_normalized_title_year"]["duplicate_groups"] = [
            {"value": value, "rows": rows}
            for value, rows in sorted(title_year_groups.items())
            if len(rows) > 1
        ]

    return result


def render_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    output = [
        "| " + " | ".join(markdown_escape(header) for header in headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        output.append("| " + " | ".join(markdown_escape(value) for value in row) + " |")
    return output


def inspect_sheet(worksheet: Any) -> tuple[list[str], dict[str, Any]]:
    max_row = worksheet.max_row or 0
    max_column = worksheet.max_column or 0
    rows_iterator = worksheet.iter_rows(values_only=True)
    raw_header = list(next(rows_iterator, ()))
    max_column = max(max_column, len(raw_header))
    if len(raw_header) < max_column:
        raw_header.extend([None] * (max_column - len(raw_header)))
    headers = [
        make_header(raw_header[column_index - 1], column_index)
        for column_index in range(1, max_column + 1)
    ]

    normalized_headers = [header["normalized"] for header in headers if header["normalized"]]
    duplicate_header_names = sorted(
        name for name, count in Counter(normalized_headers).items() if count > 1
    )
    empty_headers = [header for header in headers if not header["header"]]

    column_values: dict[int, list[Any]] = {header["index"]: [] for header in headers}
    useful_rows: list[dict[str, Any]] = []

    for row_index, row in enumerate(rows_iterator, start=2):
        values = list(row)
        if len(values) < max_column:
            values.extend([None] * (max_column - len(values)))
        values = values[:max_column]

        cells = []
        for header, value in zip(headers, values):
            column_values[header["index"]].append(value)
            cells.append(
                {
                    "letter": header["letter"],
                    "header": header["label"],
                    "value": clean_text(value),
                }
            )
        if all(is_empty(value) for value in values):
            continue
        useful_rows.append({"excel_row": row_index, "cells": cells})

    row_count = max(0, max_row - 1)
    useful_row_count = len(useful_rows)

    columns_summary = []
    for header in headers:
        values = column_values[header["index"]]
        empty_count = sum(is_empty(value) for value in values)
        normalized_non_empty_values = {
            clean_text(value) for value in values if not is_empty(value)
        }
        columns_summary.append(
            {
                "letter": header["letter"],
                "header": header["label"],
                "empty_header": not bool(header["header"]),
                "empty_values": empty_count,
                "unique_values": len(normalized_non_empty_values),
                "inferred_type": infer_column_type(values),
            }
        )

    candidates = find_candidate_columns(headers, column_values)
    duplicates = detect_duplicates(useful_rows, candidates)

    report: list[str] = []
    report.append(f"## Sheet: `{worksheet.title}`")
    report.append("")
    report.append(f"- Rows: {row_count}")
    report.append(f"- Useful rows: {useful_row_count}")
    report.append(f"- Columns: {max_column}")
    report.append("")

    report.append("### Headers")
    report.extend(
        render_table(
            ["#", "Column", "Header"],
            [[header["index"], header["letter"], header["label"]] for header in headers],
        )
    )
    report.append("")

    report.append("### Empty Or Duplicate Headers")
    if empty_headers:
        report.append(
            "- Empty headers: "
            + ", ".join(f"{header['letter']}" for header in empty_headers)
        )
    else:
        report.append("- Empty headers: none")
    if duplicate_header_names:
        report.append("- Duplicate normalized headers: " + ", ".join(duplicate_header_names))
    else:
        report.append("- Duplicate normalized headers: none")
    report.append("")

    report.append("### First Useful Rows")
    sample_rows = useful_rows[:FIRST_USEFUL_ROWS]
    if sample_rows:
        sample_headers = ["Excel row"] + [header["label"] for header in headers]
        sample_table_rows = [
            [row["excel_row"]] + [cell["value"] for cell in row["cells"]]
            for row in sample_rows
        ]
        report.extend(render_table(sample_headers, sample_table_rows))
    else:
        report.append("No useful data rows found.")
    report.append("")

    report.append("### Column Profile")
    report.extend(
        render_table(
            [
                "Column",
                "Header",
                "Empty values",
                "Unique values",
                "Inferred type",
            ],
            [
                [
                    column["letter"],
                    column["header"],
                    column["empty_values"],
                    column["unique_values"],
                    column["inferred_type"],
                ]
                for column in columns_summary
            ],
        )
    )
    report.append("")

    report.append("### Candidate Columns")
    report.extend(
        render_table(
            ["Role", "Best candidates"],
            [
                [
                    ROLE_LABELS.get(role, role),
                    ", ".join(
                        f"{item['letter']} `{item['column']}` ({item['score']})"
                        for item in matches
                    )
                    or "none",
                ]
                for role, matches in candidates.items()
            ],
        )
    )
    report.append("")

    report.append("### Possible Duplicates")
    url_duplicates = duplicates["by_letterboxd_url"]["duplicate_groups"]
    title_year_duplicates = duplicates["by_normalized_title_year"]["duplicate_groups"]
    report.append(
        f"- By Letterboxd URL: {len(url_duplicates)} duplicate group(s)"
        + (
            f" using column {duplicates['by_letterboxd_url']['candidate_column']}"
            if duplicates["by_letterboxd_url"]["candidate_column"]
            else " (no candidate column)"
        )
    )
    report.append(
        f"- By normalized title + year: {len(title_year_duplicates)} duplicate group(s)"
        + (
            " using columns "
            f"{duplicates['by_normalized_title_year']['candidate_columns']}"
            if all(duplicates["by_normalized_title_year"]["candidate_columns"].values())
            else " (missing title or year candidate)"
        )
    )
    for label, groups in [
        ("URL", url_duplicates[:20]),
        ("Title/year", title_year_duplicates[:20]),
    ]:
        if groups:
            report.append("")
            report.append(f"First duplicate groups by {label}:")
            report.extend(
                render_table(
                    ["Value", "Excel rows"],
                    [[group["value"], ", ".join(map(str, group["rows"]))] for group in groups],
                )
            )
    report.append("")

    summary = {
        "sheet_name": worksheet.title,
        "rows": row_count,
        "useful_rows": useful_row_count,
        "columns": max_column,
        "headers": headers,
        "empty_headers": [header["letter"] for header in empty_headers],
        "duplicate_normalized_headers": duplicate_header_names,
        "column_profile": columns_summary,
        "candidate_columns": candidates,
        "duplicates": duplicates,
    }
    return report, summary


def inspect_megabank() -> None:
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(f"Source file not found: {SOURCE_PATH}")

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames

    report: list[str] = [
        "# Megabank Inspection Report",
        "",
        f"- Source file: `{SOURCE_PATH.relative_to(PROJECT_ROOT)}`",
        "- Source file modified: no",
        f"- Sheets: {len(sheet_names)}",
        "",
        "## Available Sheets",
        "",
    ]
    report.extend(f"{index}. `{name}`" for index, name in enumerate(sheet_names, start=1))
    report.append("")

    json_summary: dict[str, Any] = {
        "source_file": str(SOURCE_PATH.relative_to(PROJECT_ROOT)),
        "source_file_modified": False,
        "sheets": [],
    }

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        sheet_report, sheet_summary = inspect_sheet(worksheet)
        report.extend(sheet_report)
        json_summary["sheets"].append(sheet_summary)

    REPORT_PATH.write_text("\n".join(report), encoding="utf-8")
    JSON_PATH.write_text(
        json.dumps(json_summary, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print("Megabank inspection complete.")
    print(f"Markdown report generated at: {REPORT_PATH}")
    print(f"JSON summary generated at: {JSON_PATH}")


if __name__ == "__main__":
    inspect_megabank()
